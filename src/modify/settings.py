# ============================================================================================================================
# tomodify.py - File containg all functions that don't need to be modified, are not used for plotting, nor the GUI, not the BCA
# ============================================================================================================================
# External library imports
from openpyxl.styles import PatternFill, Font, Side
from typing import Callable

# ============================================================================================================================
# Internal library imports
from methods.bv_method import bv_method
from methods.elena_method import elena_method
from methods.imv_method import imv_method
from methods.parkwind_method import parkwind_method
from modify.plots import elena_plot, plot_dop, plot_soc

# ============================================================================================================================
# Constants used throughout the program that you may 'freely' modify (at your own risk)

GUI_CONFIG: dict[str, str | int] = {
    "window_title": "FLASC Business Case Tool",  # the name that the GUI window will have
    "window_size": "900x800",  # size of the window (the window isn't resizable!)
    "max_file_name_display": 40,  # the maximum length to show of the chosen file's name
    "file_name_truncate_start": 20,  # The first character to truncate from the displayed name of the file if > 40
    "file_name_truncate_end": 15,  # the position from the last character to stop truncating
}

AVAILABLE_PLOTS: dict[str, Callable[..., None]] = (
    {  # Here lies all defined plots, add more if desired
        "State-Of-Charge": plot_soc,
        "Distribution-Of-Power": plot_dop,
        "Elena Plot":elena_plot
    }
)

INPUT_FIELDS: list[str] = (
    [  # Here lies all the entry-based user inputs, add more if desired
        "Settlement Period",
        "Storage RTE",
        "Green-Certificate Price",
        "Power Unit CAPEX",
        "Capacity Unit CAPEX",
        "Annual OPEX Rate",
        "Project Life",
        "Discount Rate",
        "Export Transmission Capacity",
        "Scenario(s) (seperate with ',' or write 'All')",
        "Timeseries Sheet Name",
        "Param Analysis Sheet Name",
        "Elena's Input",
    ]
)

STRING_BASED: list[str] = (
    [  # The names of all the entries that should be exclusively made up of a valid string, please add to this if you have added a string based entry above
        "Scenario(s) (seperate with ',' or write 'All')",
        "Timeseries Sheet Name",
        "Param Analysis Sheet Name",
        "Elena's Input",
    ]
)

CHOICE_MATRIX: dict[str, list[str]] = (
    {  # Defines the different case types, and for each case type the available methods, add if desired
        "Pure Wind": ["IMV METHOD", "BV METHOD"], # methods 0 and 1 
        "Mixed Wind and Solar": ["PARKWIND METHOD"], # method 2
        "Imagine Case": ["Elena Method"],
    }
)

METHOD_SET: list[Callable[..., None]] =(
    [
        imv_method, # method 0
        bv_method, # method 1
        parkwind_method, # method 2
        elena_method, 
    ]
)


# __________________________________________________________________________________________________________________________________________
# Excel styling constants

# Color fills for excel formatting
COLOR_FILLS: dict[str, PatternFill] = {
    "fill_purple": PatternFill(
        start_color="F2CEEF", end_color="F2CEEF", fill_type="solid"
    ),  # Light purple
    "fill_cyan": PatternFill(
        start_color="CAEDFB", end_color="CAEDFB", fill_type="solid"
    ),  # Cyan
    "fill_brown": PatternFill(
        start_color="FBE2D5", end_color="FBE2D5", fill_type="solid"
    ),  # Light brown
    "fill_green_light": PatternFill(
        start_color="DAF2D0", end_color="DAF2D0", fill_type="solid"
    ),  # Light green
    "fill_green_mid": PatternFill(
        start_color="B5E6A2", end_color="B5E6A2", fill_type="solid"
    ),  # Medium green
    "fill_green_dark": PatternFill(
        start_color="548235", end_color="548235", fill_type="solid"
    ),  # Dark green
    "fill_green_darker": PatternFill(
        start_color="8ED973", end_color="8ED973", fill_type="solid"
    ),  # Darkest green
}

# Font styles for excel formatting
BOLD_FONT = Font(bold=True)

# Border styles for excel formatting
THIN_BORDER = Side(border_style="thin", color="000000")
