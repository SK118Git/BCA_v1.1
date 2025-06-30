# ============================================================================================================================
# excel.py - File containg all functions that directly affect excel sheets 
# ============================================================================================================================
# External Imports 
import subprocess
import os
import time
import platform
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter

# ============================================================================================================================
# Internal Imports 
from frontend.popup import Progress_Popup
from libs.logger import log_print
from modify.settings import BOLD_FONT, COLOR_FILLS, THIN_BORDER

#_____________________________________________________________________________________________________________________________

#_________________________________________________________________________________________________________________________________________
# These functions define the behaviour for saving, styling and formatting to an excel sheet 

def save_to_excel(file_name:str, sheet_name:str, debug_mode:bool, data:pd.DataFrame, progress_pp:Progress_Popup) -> None:
    """ 
    Function purpose: Saves the result to excel 
    Args:
        file_name: the name of the excel file to save to 
        sheet_name: the name of the excel sheet to save to 
        debug_mode: if True adds some extra print statements in the code to help backtracing and debugging 
        data: the result dataframe to save 
        progress_pp: a class which contains the progress bar and its label 
    """

    log_print(f"\nWriting to {file_name} on sheet called {sheet_name}.\n ")

    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists= 'replace') as writer: 
        data.to_excel(writer,sheet_name=sheet_name,index=False )

    progress_pp.update_vals("Successfully saved to excel, now styling.", 33.33)
  


    style_excel_sheet(file_name, sheet_name, debug_mode)
    progress_pp.update_vals("Successfully styled, now formatting.", 66.66)


    format_excel_sheet(file_name, sheet_name, debug_mode)
    progress_pp.update_vals("Successfully formatted! \n Program Execution Complete.", 100.00)
    

    return 



def style_excel_sheet(file_name:str, sheet_name:str, debug_mode:bool) -> None:
    """ 
    Function purpose: This function does the styling for the excel sheet (the colors, column size, borders) \n
    Args: 
        file_path: the name of the excel file \n
        sheet_name: the name of the sheet that is being formatted \n
        debug_mode:  if True adds some extra print statements in the code to help backtracing and debugging \n
    """

    log_print("Styling... \n ")

    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Styling definitions
    thin_border = THIN_BORDER
    pink_fill = COLOR_FILLS['fill_purple'] 
    cyan_fill = COLOR_FILLS['fill_cyan'] 
    salmon_fill = COLOR_FILLS['fill_brown'] 
    light_green_fill = COLOR_FILLS['fill_green_light'] 
    dark_green_fill = COLOR_FILLS['fill_green_dark'] 

    headers = [cell.value for cell in ws[1]]
    header_index = {str(h).strip().lower(): idx + 1 for idx, h in enumerate(headers)}
    max_row = ws.max_row
    max_col = len(headers)

    def get_col(name):
        val = header_index.get(name.lower().strip())
        if val:
            return val 
        else:
            raise ValueError("Invalid Column name")

    def add_border(cell, side_name, border_side):
        """Safely add a border to a single side without erasing others."""
        current = cell.border
        cell.border = Border(
            left=border_side if side_name == "left" else current.left,
            right=border_side if side_name == "right" else current.right,
            top=border_side if side_name == "top" else current.top,
            bottom=border_side if side_name == "bottom" else current.bottom
        )
        return 

    col_duration = get_col("duration")
    col_storage_capex = get_col("storage capex")
    col_storage_opex = get_col("storage opex")
    col_irr = get_col("irr")
    col_npv = get_col("npv")

    ### 1. Color header cells
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        cell = ws[f"{col_letter}1"]

        if col <= col_duration:
            cell.fill = pink_fill
        elif col_storage_capex and col == col_storage_capex:
            cell.fill = salmon_fill
        elif col_storage_opex and col == col_storage_opex:
            cell.fill = salmon_fill
        elif col_irr and col == col_irr:
            cell.fill = dark_green_fill
        elif col_npv and col == col_npv:
            cell.fill = dark_green_fill
        elif col_duration < col < col_storage_capex:
            cell.fill = cyan_fill
        elif col_storage_opex < col < col_irr:
            cell.fill = light_green_fill

    ### 2. Horizontal line between row 1 and 2 and make first row bold
    for col in range(1, max_col + 1):
        temp_cell = ws.cell(row=2, column=col)
        add_border(temp_cell, "top", thin_border)
        ws.cell(row=1, column=col).font = BOLD_FONT


    ### 3. Vertical line after Duration
    if col_duration:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_duration) 
            add_border(temp_cell, "right", thin_border)

    ### 4. Vertical line left of Storage Capex
    if col_storage_capex:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_storage_capex)
            add_border(temp_cell, "left", thin_border)

    ### 5. Vertical line right of Storage Opex
    if col_storage_opex:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_storage_opex)
            add_border(temp_cell, "right", thin_border)

    ### 6. Vertical line left of IRR
    if col_irr:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_irr)
            add_border(temp_cell, "left", thin_border)

    ### 7. Vertical line right of NPV
    if col_npv:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_npv)
            add_border(temp_cell, "right", thin_border)

    ### 8. Horizontal line on letter change in column A (extended across to NPV)
    prev_letter = None
    for row in range(2, max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and val:
            first_letter = val[0].lower()
            if prev_letter is not None and first_letter != prev_letter:
                for col in range(1, col_npv + 1):
                    temp_cell = ws.cell(row=row, column=col)
                    add_border(temp_cell, "top", thin_border)
            prev_letter = first_letter

    ### 9. Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue
        if col_idx == col_irr:
            width = len(str(header)) * 3
        elif col_idx == col_npv:
            width = len(str(header)) * 5
        else:
            width = len(str(header)) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(file_name)

    log_print("Done styling.")
    return 


def format_excel_sheet(file_name:str, sheet_name:str, debug_mode:bool):
    """ 
    Function purpose: This function formats the values of the excel sheet \n
    Args: 
        file_name: the name of the excel file 
        sheet_name: the name of the sheet that is being saved 
        debug_mode: if True adds some extra print statements in the code to help backtracing and debugging 
    """ 

    log_print("Formatting columns: \n ")
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Mapping of columns to formatting codes
    custom_formats = {
        'ppa price': '#,##0.0',
        'wind power (mw)': '#,##0.00" MW"',  
        'solar installed (mwp)': '#,##0.00" MW"',  
        'balancing market participation': '0%',
        'storage power rating': '0.000" MW"',
        'duration': '0.00" hours"',
        'irr': '0.00%',
        'npv': '#,##0.0" €"'
    }


    header = [cell.value for cell in ws[1]]
    for col_idx, col_name in enumerate(header, start=1):
        col_name_lc = col_name.lower() if isinstance(col_name, str) else ""
        
        if col_name_lc in custom_formats:
            log_print(f"\t Doing: {col_name_lc}.")
            fmt = custom_formats[col_name_lc]
        elif col_name:  # all other named columns
            fmt = '#,##0.0'
        else:
            continue  # skip unnamed columns

        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = fmt
    log_print("Done Formatting!")
    wb.save(file_name)
    return 




#__________________________________________________________________________________________________________________________________
# This code defines a function which eliminates many bugs 

def force_excel_calc(file_name:str) -> bool:
    """
    Function purpose: Force Excel to recalculate all formulas. \n
    Outputs: Whether or not the operation was successful as a boolean 
    Note: On macOS, uses AppleScript only to trigger the automation permission prompt, then uses xlwings for actual recalculation.
    Args: 
        file_name: the name of the excel file 
    
    """
    is_macos: bool = platform.system() == 'Darwin'

    if is_macos:
        log_print("Running on macOS...")

        abs_path = os.path.abspath(file_name)

        # This triggers the Automation permission prompt
        log_print("Triggering macOS Automation prompt...")
        try:
            trigger_script = f'tell application "Microsoft Excel" to get name of active workbook'
            result = subprocess.run(['osascript', '-e', trigger_script], capture_output=True, text=True)
            log_print(f"osascript result: {result.stderr.strip()}")
        except Exception as e:
            log_print(f"Error while triggering permission: {e}")

        time.sleep(2)

        # Now try xlwings (should work if permission was granted)
        try:
            log_print("Using xlwings to recalculate Excel formulas...")
            app = xw.App(visible=False)
            wb = app.books.open(file_name)
            wb.app.calculate()
            wb.save()
            wb.close()
            app.quit()
            log_print(f"Recalculation complete: {file_name}.")
            return True

        except Exception as e:
            log_print(f"Alert, xlwings error: {e}")
            if "-1743" in str(e) or "not authorized" in str(e).lower():
                log_print("Automation permission denied.")
                log_print("Go to System Settings -> Privacy & Security -> Automation.")
                log_print("Enable control of Microsoft Excel for your app.")
            return False

    else:
        log_print("Running on non-macOS system...")
        try:
            app = xw.App(visible=False)
            wb = app.books.open(file_name)
            wb.app.calculate()
            wb.save()
            wb.close()
            app.quit()
            log_print(f"Successfully recalculated: {file_name}.")
            return True
        except Exception as e:
            log_print(f"Alert, xlwings error: {e}")
            return False
        
    # If we get here, everything failed
    log_print("All automated methods failed.")
    log_print("Please manually open the Excel file, press Ctrl+Shift+F9 to recalculate, then save.")
    return False
