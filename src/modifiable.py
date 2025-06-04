# ============================================================================================================================
# tomodify.py - File containg all functions that don't need to be modified, are not used for plotting, nor the GUI, not the BCA 
# ============================================================================================================================
# External library imports 
import numpy as np 
import pandas as pd
from openpyxl.styles import PatternFill, Font, Side
from typing import Callable
# ============================================================================================================================
# Internal library imports
from plots import plot_dop, plot_soc
# ============================================================================================================================
# Constants used throughout the program that you may 'freely' modify (at your own risk)

GUI_CONFIG: dict[str, str|int] = {
    'window_title': 'FLASC Business Case Tool',     # the name that the GUI window will have 
    'window_size': '800x700',                       # size of the window (the window isn't resizable!)
    'max_file_name_display': 40,                     # the maximum length to show of the chosen file's name 
    'file_name_truncate_start': 20,                  # The first character to truncate from the displayed name of the file if > 40 
    'file_name_truncate_end': 15                     # the position from the last character to stop truncating 
}

AVAILABLE_PLOTS: dict[str, Callable[..., None]] = {                                 # Here lies all defined plots, add more if desired
    "State-Of-Charge": plot_soc,                             
    "Distribution-Of-Power": plot_dop, 
}

INPUT_FIELDS: list[str] = [                                    # Here lies all the entry-based user inputs, add more if desired
    "Settlement Period",
    "Storage RTE",
    "Green-Certificate Price",
    "Power Unit CAPEX",
    "Capacity Unit CAPEX", 
    "Annual OPEX Rate",
    "Project Life",
    "Discount Rate",
    "Export Transmission Capacity", 
    "Scenario(s) (seperate with ',' or write 'All')",
    "Timeseries Sheet Name",
    "Param Analysis Sheet Name"
]

STRING_BASED: list[str] = [                                    # The names of all the entries that should be exclusively made up of a valid string, please add to this if you have added a string based entry above 
    "Scenario(s) (seperate with ',' or write 'All')",
    "Timeseries Sheet Name",
    "Param Analysis Sheet Name"
]

CHOICE_MATRIX: dict[str, list[str]]= {                                    # Defines the different case types, and for each case type the available methods, add if desired 
    "Pure Wind":["IMV Method", "BV METHOD"],
    "Mixed Wind and Solar": ["PARKWIND METHOD"],
    "Pure Solar": ["TO BE IMPLEMENTED"]
}

#__________________________________________________________________________________________________________________________________________
# Excel styling constants

# Color fills for excel formatting
COLOR_FILLS: dict[str, PatternFill] = {
    "fill_purple": PatternFill(start_color="F2CEEF", end_color="F2CEEF", fill_type="solid"),        # Light purple
    "fill_cyan" : PatternFill(start_color="CAEDFB", end_color="CAEDFB", fill_type="solid"),          # Cyan
    "fill_brown" : PatternFill(start_color="FBE2D5", end_color="FBE2D5", fill_type="solid"),         # Light brown
    "fill_green_light" : PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid"),   # Light green
    "fill_green_mid" : PatternFill(start_color="B5E6A2", end_color="B5E6A2", fill_type="solid"),     # Medium green
    "fill_green_dark" : PatternFill(start_color="548235", end_color="548235", fill_type="solid"),    # Dark green
    "fill_green_darker" : PatternFill(start_color="8ED973", end_color="8ED973", fill_type="solid"),  # Darkest green
}

# Font styles for excel formatting
BOLD_FONT = Font(bold=True)

# Border styles for excel formatting
THIN_BORDER = Side(border_style="thin", color="000000")

# ============================================================================================================================

def calculate_ap(df:pd.DataFrame, method:int) -> pd.Series:
    """
    Function purpose: Calculated the Available Power differently depending on the chosen method \n 
    Inputs: \n
        df = a pandas Dataframe containing the timeseries sheet's values \n
        method = the chosen method \n
    Outputs: the column containing the values of Available Transmission Capacity that was computed \n
    Notes:
    """
    if method == 1: #BV
        df['Available Power [MW]'] = df['Potential Generation [MW]'] - df['Generation Constraint [MW]']
        df['Available Power [MW]'] = np.where(df['Available Power [MW]'] < 0, 0, df['Available Power [MW]'])
    elif method == 0: #IMV
        def siemens_gamesa_power_curve(wind_speed:float)-> float:
            """
            Approximate power curve for Siemens Gamesa SG 14-222 DD wind turbine.
            Input: Wind speed (m/s)
            Output: Power output (MW)
            """
            cut_in = 3       # Minimum wind speed for power generation (m/s)
            rated = 13       # Wind speed where full power is reached (m/s)
            cut_out = 32     # Wind speed where turbine shuts down (m/s)
            max_power = 14   # Rated power output (MW)

            if wind_speed < cut_in or wind_speed >= cut_out:
                return 0  # No power generation
            
            elif wind_speed >= rated:
                return max_power  # Full rated power

            else:
                # Use a logistic (sigmoid) function to approximate smooth ramp-up
                return max_power / (1 + np.exp(-0.5 * (wind_speed - (cut_in + rated) / 2)))

        num_turbines = 72 #1000MW / 14MW = 71.42

        # wake_loss = 0.1 #Wake loss fraction (typically 5-15% offshore)
        # blockage_loss = 0.02 #Blockage loss fraction (typically 1-3% offshore)

        #df['Available Power [MW]'] = (1-wake_loss)*(1-blockage_loss)*num_turbines*df['Wind Speed [m/s]'].apply(siemens_gamesa_power_curve)
        df['Available Power [MW]'] = num_turbines*df['Wind Speed [m/s]'].apply(siemens_gamesa_power_curve)
    return df['Available Power [MW]']

#__________________________________________________________________________________________________________________________________________
def calculate_atc(df:pd.DataFrame, method:int, input_values:dict[str, float|str]) -> pd.Series:
    """
    Function purpose: Calculated the Available Power differently depending on the chosen method \n 
    Inputs: \n
        df = a pandas Dataframe containing the timeseries sheet's values \n
        method = the chosen method \n
        input_values = a dictionnary containing all the values inputted by the user \n
    Outputs: the column containing the values of Available Power that was computed \n
    Notes:
    """
    # Replace with your actual project name
    if method==0: #imv-like
        transformer_rating = 1000 #MW
        #maximum power
        df['Available Transmission Capacity [MW]'] = transformer_rating - df['Capacity Constraint [MW]'] #transmission power rating
    

    elif method == 1: #BV
        #%% Compute Transmisstion Capacity
        transformer_rating = 9.5*2 #MW

        #maximum power: "+ df['Available Power [MW]']" is included since the curtailment figures provided by the developer are relative to the Available Power not the transmission capacity: 
            #E.g. Transmission Capacity = 19MW: 
            #Available Power = 10WM
            #Curtailment = 2MW (realtive to Available Power)
            #Effective Transmission Capacity = 19 - ((19 - 10) + 2) = 8MW 
        df['Available Transmission Capacity [MW]'] = np.where(df['Capacity Constraint [MW]'] == 0,  transformer_rating,
                                                    df['Actual Generation [MW]']) #effective transmission power rating

        df['Available Transmission Capacity [MW]'] = np.where (df['Available Transmission Capacity [MW]'] < 0, 0, df['Available Transmission Capacity [MW]'])
    

    elif method == 2: #parkwind
        #%% Compute Transmisstion Capacity
        transformer_rating = input_values["Export Transmission Capacity"] #MW

        #maximum power
        df['Available Transmission Capacity [MW]'] = transformer_rating - df['Capacity Constraint [MW]'] #transmission power rating
        
    return df['Available Transmission Capacity [MW]'] 