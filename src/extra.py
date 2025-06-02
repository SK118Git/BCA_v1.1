# ============================================================================================================================
# extra.py - File containg all functions that don't need to be modified, are not used for plotting, nor the GUI, not the BCA 
# ============================================================================================================================
# External Imports 
#_____________________________________________________________________________________________________________________________
# Main imports 
import pandas as pd 
import numpy as np
import numpy_financial as npf

#_____________________________________________________________________________________________________________________________
# Imports required for manipulating the excel sheet 
from openpyxl import load_workbook
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter

#_____________________________________________________________________________________________________________________________
# This library, and the associated force_excel_calc function serve to force excel to recalculate all values in the sheets as 
# otherwise any cell defined by a formula would not be read by python 
import xlwings as xw

# ============================================================================================================================
# Internal Imports (The required constants) 
from tomodfiy import BOLD_FONT, THIN_BORDER, COLOR_ZONES, COLOR_ZONES_LONG
# ============================================================================================================================


# ============================================================================================================================
# Functions that define methods or utilities for dictionnaries 

def find_scenario_index(df, scenario):
    """ 
    Function purpose: Finds the row number of a given scenario to then be able to index the other values by that number 
    Inputs: 
        df = the dataframe (or excel sheet) where all the scenarios are stored with their respective parameters
        scenario = the label of the scenario (ex: B7)
    Outputs: the row number - 1 where the given scenario label is located on the excel sheet 
    """
    index_value = df.index[df["Scenario"] == scenario].tolist()
    return index_value[0] if index_value else None  # Returns None if not found


def update_dict(my_dict, key, new_value):
    """ 
    Function purpose: Updates 
    Inputs: Takes a python dictionnary and add a value to a key
        my_dict = any python dictionnary where each key is associated to a LIST 
        key = the key in the dictionnary where the entry needs to be added
        new_value = the new value to add 
    Outputs: None
    """
    if key not in my_dict: 
    # if the key doesn't exist yet then create it and directly assign the value
        my_dict[key] = [new_value]
    elif len(my_dict[key]) > 1:
    # if the key has at least 2 values in its associated list, replace the second value
        my_dict[key][1] = new_value
    else:
    # else add the new value to the list associated to the key 
         my_dict[key].append(new_value)



def find_index(container, search_for):
    """ 
    Function purpose: Finds the index of a given key in a dictionnary  
    Inputs: 
        container = a python dictionnary
        search_for = the key to find
    Outputs: the index of the key, or -1 if an error has occurred
    """
    try:
        return list(container.keys()).index(search_for)
    except ValueError:
        return -1  # Key not found
    

# ============================================================================================================================
# All tht follows are functions useful for the BCA 
def safe_irr(cash_flows):
    """ 
    Function purpose:  Ensures the calculated irr value isn't a completely unrealistic value
    Inputs: the cash flows
    Outputs: a cleaned irr 
    """
    irr = npf.irr(cash_flows)
    
    if np.isnan(irr):
        # Case where IRR is not computable
        return -0.20
    elif irr > 1.0:
        # Case where IRR is unrealistically high (>100%)
        return np.nan
    else:
        return irr


# ============================================================================================================================
# All that follows are useful in the managing of excel files:

#_____________________________________________________________________________________________________________________________
# These functions are used for managing pandas Dataframes read from Excel files

def read_df(filename, df_sheetname):
    """ 
    Function purpose: Reads and turns into a pandas Dataframe the timeseries sheet
    Inputs:
        filename = the name of the excel file
        df_sheetname = name of the Timeseries sheet
    Outputs: pandas Dataframe of said excel sheet 
    """
    df = pd.read_excel(filename, sheet_name=df_sheetname, header=0, engine="openpyxl")
    return df


def read_pdf(filename, pdf_sheetname):
    """ 
    Function purpose:  Reads the excel sheet where the parameters by scenario are located 
    Inputs: 
        filename = the name of the excel file 
        pdf_sheetname = the name of the excel sheet where the parameters are defiend for each scenario
    Outputs: a pandas Dataframe of said excel sheet 
    Note: the reason this function is different than the other is that in the original code 
          (probably to make the calculations doable for numpy) the dataframe was manipulated in
          a particular way. This code slightly modified the previous method to fix a few bugs that arose. 
    """

     # Read Input Parameters from the param_df
    wb = load_workbook(filename, data_only=True)
    sheet = wb[pdf_sheetname]

    # Convert sheet to DataFrame manually
    data = sheet.values
    cols = next(data)   
    param_df = pd.DataFrame(data, columns=cols)
    #param_df = pd.read_excel(filename, sheet_name="Parametric Analysis", header=0, engine="openpyxl")

    return param_df 

#_________________________________________________________________________________________________________________________________________
# These functions define the behaviour for saving, styling and formatting to an excel sheet 

def save_to_excel(filename, data, sheet_name, case_type, debug_mode):
    """ 
    Function purpose: Saves the result to excel 
    Inputs: 
        filename = the name of the excel file to save to 
        data = the result dataframe to save 
        sheet_name = the name of the excel sheet to save to
        case_type = the type of case that is being studied (this modifies the formatting rule so is a necessary parameter)
        debug_mode = if True adds some extra print statements in the code to help backtracing and debugging
    Outputs: None
    """
    print(f"\nWriting to {filename} on sheet called {sheet_name}... \n ")
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists= 'replace') as writer: 
        data.to_excel(writer,sheet_name=sheet_name,index=False )

    print("\nFormatting... \n ")
    style_excel(filename, sheet_name, case_type)
    format_excel(filename, sheet_name, case_type, debug_mode)
    print("\n Done Formatting! \n ")
    return 

def style_excel(file_path, sheet_name, case_type):
    """ 
    Function purpose: This function does the styling for the excel sheet (the colors, column size, borders)
    Inputs: 
        file_path: the name of the excel file 
        sheet_name = the name of the sheet that is being formatted
        case_type = the type of case that is being studied 
    Outputs:  None
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    bold_font = BOLD_FONT
    thin_border = THIN_BORDER

    # Define color zones
    if case_type == 0:
        color_zones = COLOR_ZONES 
    else: 
        color_zones = COLOR_ZONES_LONG

    # Fill first row and auto-size columns
    for start_col, end_col, fill in color_zones:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = max(10, len(str(cell.value)) + 2)
    

    # Bold first column (A)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold_font

    # Add vertical border between column A and B onward (right edge of col 1)
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.border = Border(right=thin_border)

    # Add vertical border between each color zone
    for color in color_zones:
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=color[1])
            cell.border = Border(right=thin_border)


    # Draw repeating 10-row-high boxes based on color zones
    for start_row in range(2, 102, 10):
        end_row = start_row + 9
        for start_col, end_col, _ in color_zones:
            for row in range(start_row, min(end_row + 1, ws.max_row + 1)):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = Border(
                        top=thin_border if row == start_row else None,
                        bottom=thin_border if row == end_row else None,
                        left=thin_border if col == start_col else None,
                        right=thin_border if col == end_col else None
                    )

    # --- Save workbook ---
    wb.save(file_path)
    return


def format_excel(filename, sheet_name, case_type, debug_mode):
    """ 
    Function purpose: This function formats the values of the excel sheet
    Inputs: 
        filename = the name of the excel file 
        sheet_name = the name of the sheet that is being saved 
        case_type = the type of case that is being studied
        debug_mode = if True enables additional print statements to help backtrace and debug 
    Outputs: None
    """ 
    if debug_mode: print("Entered formatting")
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    
    # Get max row
    max_row = ws.max_row
    
    if case_type == 1:
        color_zones = COLOR_ZONES_LONG
    else:
        color_zones = COLOR_ZONES
  
    
    # Extract column indices from the color zones
    zone_first = color_zones[0]      # First zone (A-G or A-H)
    zone_second = color_zones[1]     # Second zone (H-M or I-N)
    zone_third = color_zones[2]      # Third zone (N-O or O-P)
    zone_fourth = color_zones[3]     # Fourth zone (P or Q)
    zone_fifth = color_zones[4]      # Fifth zone (Q-T or R-U)
    zone_sixth = color_zones[5]      # Sixth zone (U-V or V-W)
    
    # Determine column positions based on zone boundaries
    # For percentage column: 2nd column before end of first zone
    col_percentage_1_idx = zone_first[1] - 2  # E in normal, F in long
    # For hours column: end of first zone
    col_hours_idx = zone_first[1]  # G in normal, H in long
    # For percentage column 2: first column of last zone
    col_percentage_2_idx = zone_sixth[0]  # U in normal, V in long
    # For euro column: second column of last zone
    col_euro_idx = zone_sixth[1]  # V in normal, W in long
    
    # Range for thousand separator formatting: from second zone start to fifth zone end
    col_range_thousands = list(range(zone_second[0], zone_fifth[1] + 1))
    
    for row in range(2, max_row + 1):  # Skip header (row 1)
        # First percentage column (E in normal, F in long)
        cell = ws.cell(row=row, column=col_percentage_1_idx)
        cell.number_format = '0%'
        
        # Hours column (G in normal, H in long): append " hours" (convert to string)
        cell = ws.cell(row=row, column=col_hours_idx)
        val = cell.value
        if isinstance(val, (int, float)):
            cell.value = f"{int(round(val))} hrs"
            cell.number_format = '@'  # Text format
        elif isinstance(val, str) and val.strip().isdigit():
            cell.value = f"{int(val.strip())} hrs"
            cell.number_format = '@'
        
        # Thousand separator columns: thousand separator, rounded integer
        for col_idx in col_range_thousands:
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.value = round(cell.value)
                cell.number_format = '#,##0'
        
        # Second percentage column (U in normal, V in long)
        cell = ws.cell(row=row, column=col_percentage_2_idx)
        cell.number_format = '#.0%'
        
        # Euro column (V in normal, W in long): Euro, thousand separator, rounded integer
        cell = ws.cell(row=row, column=col_euro_idx)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.value = round(cell.value)
            cell.number_format = '€#,##0'
    
    # Save workbook
    save_path = filename 
    wb.save(save_path)
    if debug_mode:
        print("Sheet names:", wb.sheetnames)
        print("Using sheet:", sheet_name)
        print(f"Saved to: {save_path}")

        print("Done")
    return


#__________________________________________________________________________________________________________________________________
# This code defines a function which eliminates many bugs 

def force_excel_calc(filename):
    """ 
    Function purpose: This function forces excel to recalculate all the values in a given file
    Inputs: 
        filename = the name of the excel file 
    Outputs: None
    Note: Without this function python won't properlu read any cells defined by an expression ("= ...")
          and as such is crucial. 
          Without this function the code has endless problems as many key cells will be just read as empty
    """ 
    app = xw.apps.active
    if app is None:
        app = xw.App(visible=False)
    wb = app.books.open(filename)
    wb.app.calculate()
    wb.save()
    wb.close()
    app.quit()
    return 


