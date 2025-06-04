# ============================================================================================================================
# excel.py - File containg all functions that directly affect excel sheets 
# ============================================================================================================================
# External Imports 
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter
from typing import Optional
from tkinter import Label 
from tkinter.ttk import Progressbar

# ============================================================================================================================
# Internal Imports (The required constants) 
from modifiable import BOLD_FONT, COLOR_FILLS, THIN_BORDER

#_____________________________________________________________________________________________________________________________
# These functions are used for managing pandas Dataframes read from Excel files

def read_df(file_name, df_sheetname) -> pd.DataFrame:
    """ 
    Function purpose: Reads and turns into a pandas Dataframe the timeseries sheet \n
    Inputs: \n
        file_name = the name of the excel file \n
        df_sheetname = name of the Timeseries sheet \n
    Outputs: pandas Dataframe of said excel sheet 
    """
    df:pd.DataFrame = pd.read_excel(file_name, sheet_name=df_sheetname, header=0, engine="openpyxl")
    return df


def read_pdf(file_name, pdf_sheetname) -> pd.DataFrame:
    """ 
    Function purpose:  Reads the excel sheet where the parameters by scenario are located \n
    Inputs: \n
        file_name = the name of the excel file \n
        pdf_sheetname = the name of the excel sheet where the parameters are defiend for each scenario \n
    Outputs: a pandas Dataframe of said excel sheet \n
    Note: the reason this function is different than the other is that in the original code 
          (probably to make the calculations doable for numpy) the dataframe was manipulated in 
          a particular way. This code slightly modified the previous method to fix a few bugs that arose. 
    """

     # Read Input Parameters from the param_df
    wb = load_workbook(file_name, data_only=True)
    sheet = wb[pdf_sheetname]

    # Convert sheet to DataFrame manually
    data = sheet.values
    cols = next(data)   
    param_df = pd.DataFrame(data, columns=cols)

    return param_df 


#_________________________________________________________________________________________________________________________________________
# These functions define the behaviour for saving, styling and formatting to an excel sheet 

def save_to_excel(file_name:str, sheet_name:str, debug_mode:bool, data:pd.DataFrame,  progress_bar:Optional[Progressbar]=None, progress_label:Optional[Label]=None) -> None:
    """ 
    Function purpose: Saves the result to excel \n 
    Inputs: \n
        file_name = the name of the excel file to save to \n
        sheet_name = the name of the excel sheet to save to \n
        debug_mode = if True adds some extra print statements in the code to help backtracing and debugging \n
        data = the result dataframe to save \n
        progres_bar = the progress bar, set to optional for compatibility with tests \n
        progress_label = the label that appears above the progress bar, set to optional for compatibility with tests \n
    Outputs: None
    """
    if debug_mode: print(f"\nWriting to {file_name} on sheet called {sheet_name}... \n ")

    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists= 'replace') as writer: 
        data.to_excel(writer,sheet_name=sheet_name,index=False )

    if progress_bar and progress_label:
            # Reset progress
            progress_bar['value'] = 33.33
            progress_bar.update_idletasks()

            # Change label text
            progress_label.config(text="Successfully saved to excel, now styling")
            progress_label.update_idletasks()


    if debug_mode: print("\nFormatting... \n ")
    style_excel_sheet(file_name, sheet_name, debug_mode)

    if progress_bar and progress_label:
            # Reset progress
            progress_bar['value'] = 66.66
            progress_bar.update_idletasks()

            # Change label text
            progress_label.config(text="Successfully styled, now formatting")
            progress_label.update_idletasks()


    format_excel_sheet(file_name, sheet_name, debug_mode)
    if progress_bar and progress_label:
            # Reset progress
            progress_bar['value'] = 100.00
            progress_bar.update_idletasks()

            # Change label text
            progress_label.config(text="Successfully formatted! \n Program Execution Complete")
            progress_label.update_idletasks()

    if debug_mode: print("\n Done Formatting! \n ")

    return 



def style_excel_sheet(file_name:str, sheet_name:str, debug_mode:bool) -> None:
    """ 
    Function purpose: This function does the styling for the excel sheet (the colors, column size, borders) \n
    Inputs: \n
        file_path: the name of the excel file \n
        sheet_name = the name of the sheet that is being formatted \n
        debug_mode = if True adds some extra print statements in the code to help backtracing and debugging \n
    Outputs:  None
    """
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Styling definitions
    thin_border = THIN_BORDER
    pink_fill = COLOR_FILLS['fill_purple'] 
    cyan_fill = COLOR_FILLS['fill_cyan'] 
    salmon_fill = COLOR_FILLS['fill_brown'] 
    light_green_fill = COLOR_FILLS['fill_green_light'] 
    dark_green_fill = COLOR_FILLS['fill_green_dark'] 

    headers = [cell.value for cell in ws[1]]
    header_index = {str(h).strip().lower(): idx + 1 for idx, h in enumerate(headers)}
    max_row = ws.max_row
    max_col = len(headers)

    def get_col(name):
        val = header_index.get(name.lower().strip())
        if val:
            return val 
        else:
            raise ValueError("Invalid Column name")

    def add_border(cell, side_name, border_side):
        """Safely add a border to a single side without erasing others."""
        current = cell.border
        cell.border = Border(
            left=border_side if side_name == "left" else current.left,
            right=border_side if side_name == "right" else current.right,
            top=border_side if side_name == "top" else current.top,
            bottom=border_side if side_name == "bottom" else current.bottom
        )
        return 

    col_duration = get_col("duration")
    col_storage_capex = get_col("storage capex")
    col_storage_opex = get_col("storage opex")
    col_irr = get_col("irr")
    col_npv = get_col("npv")

    ### 1. Color header cells
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        cell = ws[f"{col_letter}1"]

        if col <= col_duration:
            cell.fill = pink_fill
        elif col_storage_capex and col == col_storage_capex:
            cell.fill = salmon_fill
        elif col_storage_opex and col == col_storage_opex:
            cell.fill = salmon_fill
        elif col_irr and col == col_irr:
            cell.fill = dark_green_fill
        elif col_npv and col == col_npv:
            cell.fill = dark_green_fill
        elif col_duration < col < col_storage_capex:
            cell.fill = cyan_fill
        elif col_storage_opex < col < col_irr:
            cell.fill = light_green_fill

    ### 2. Horizontal line between row 1 and 2 and make first row bold
    for col in range(1, max_col + 1):
        temp_cell = ws.cell(row=2, column=col)
        add_border(temp_cell, "top", thin_border)
        ws.cell(row=1, column=col).font = BOLD_FONT


    ### 3. Vertical line after Duration
    if col_duration:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_duration) 
            add_border(temp_cell, "right", thin_border)

    ### 4. Vertical line left of Storage Capex
    if col_storage_capex:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_storage_capex)
            add_border(temp_cell, "left", thin_border)

    ### 5. Vertical line right of Storage Opex
    if col_storage_opex:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_storage_opex)
            add_border(temp_cell, "right", thin_border)

    ### 6. Vertical line left of IRR
    if col_irr:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_irr)
            add_border(temp_cell, "left", thin_border)

    ### 7. Vertical line right of NPV
    if col_npv:
        for row in range(2, max_row + 1):
            temp_cell = ws.cell(row=row, column=col_npv)
            add_border(temp_cell, "right", thin_border)

    ### 8. Horizontal line on letter change in column A (extended across to NPV)
    prev_letter = None
    for row in range(2, max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and val:
            first_letter = val[0].lower()
            if prev_letter is not None and first_letter != prev_letter:
                for col in range(1, col_npv + 1):
                    temp_cell = ws.cell(row=row, column=col)
                    add_border(temp_cell, "top", thin_border)
            prev_letter = first_letter

    ### 9. Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue
        if col_idx == col_irr:
            width = len(str(header)) * 3
        elif col_idx == col_npv:
            width = len(str(header)) * 5
        else:
            width = len(str(header)) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(file_name)

    if debug_mode: print("Done styling")
    return 


def format_excel_sheet(file_name:str, sheet_name:str, debug_mode:bool):
    """ 
    Function purpose: This function formats the values of the excel sheet \n
    Inputs: \n
        file_name = the name of the excel file \n
        sheet_name = the name of the sheet that is being saved \n
        debug_mode = if True adds some extra print statements in the code to help backtracing and debugging \n
    Outputs: None
    """ 
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Mapping of columns to formatting codes
    custom_formats = {
        'ppa price': '#,##0.0',
        'wind power (mw)': '#,##0.00" MW"',  
        'solar installed (mwp)': '#,##0.00" MW"',  
        'balancing market participation': '0%',
        'storage power rating': '0.000" MW"',
        'duration': '0.00" hours"',
        'irr': '0.00%',
        'npv': '#,##0.0" €"'
    }


    header = [cell.value for cell in ws[1]]
    for col_idx, col_name in enumerate(header, start=1):
        col_name_lc = col_name.lower() if isinstance(col_name, str) else ""
        
        if col_name_lc in custom_formats:
            if debug_mode: print(col_name_lc)
            fmt = custom_formats[col_name_lc]
        elif col_name:  # all other named columns
            fmt = '#,##0.0'
        else:
            continue  # skip unnamed columns

        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = fmt
    if debug_mode: print("Done")
    wb.save(file_name)




#__________________________________________________________________________________________________________________________________
# This code defines a function which eliminates many bugs 

def force_excel_calc(file_name:str): 
    """ 
    Function purpose: This function forces excel to recalculate all the values in a given file \n
    Inputs: \n
        file_name = the name of the excel file \n
    Outputs: None \n
    Note: Without this function python won't properlu read any cells defined by an expression ("= ...")
          and as such is crucial. \n
          Without this function the code has endless problems as many key cells will be just read as empty
    """ 
    app = xw.apps.active
    if app is None:
        app = xw.App(visible=False)
    wb = app.books.open(file_name)
    wb.app.calculate()
    wb.save()
    wb.close()
    app.quit()
    return 


